import os
import glob
import zipfile
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def obter_nome_arquivo_saida(pasta_destino):
    """
    Gera o nome do arquivo respeitando a regra:
    Implantação DD.MM.YYYY
    Se houver mais de um no dia, Implantação DD.MM.YYYY V2... e assim por diante.
    """
    hoje_str = datetime.now().strftime("%d.%m.%Y")
    base_nome = f"Implantação {hoje_str}"
    
    padrao_busca = os.path.join(pasta_destino, f"{base_nome}*.xlsx")
    arquivos_existentes = glob.glob(padrao_busca)
    
    if not arquivos_existentes:
        return f"{base_nome}.xlsx"
    
    # Se já existir, precisamos calcular o número da versão (V2, V3, etc)
    return f"{base_nome} V{len(arquivos_existentes) + 1}.xlsx"

def gerar_relatorio_final(dados_processados, estatisticas, pasta_destino_pdfs, pasta_destino_output):
    """
    dados_processados: Lista de dicionários com todos os dados montados da execução.
    estatisticas: Dict contendo start_time, end_time, sucesso, falha_extracao, senha_errada, sem_senha
    """
    if not os.path.exists(pasta_destino_output):
        os.makedirs(pasta_destino_output)

    nome_excel = obter_nome_arquivo_saida(pasta_destino_output)
    caminho_excel = os.path.join(pasta_destino_output, nome_excel)

    # Cores (Fontes)
    fonte_verde = Font(color="00B050", bold=True)
    fonte_vermelha = Font(color="FF0000", bold=True)
    fonte_padrao = Font(color="000000")

    wb = openpyxl.Workbook()
    
    # ======== ABA 1: Resumo ========
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Contar PDFs na pasta
    qtd_pdfs = len([f for f in os.listdir(pasta_destino_pdfs) if f.endswith('.pdf')])
    sucessos = estatisticas.get('sucesso', 0)
    implantacao_ou_pab = estatisticas.get('implantacao_ou_pab', 0)

    linhas_resumo = [
        ["Data:", datetime.now().strftime("%d/%m/%Y")],
        ["Horário início:", estatisticas.get('start_time', '')],
        ["Horário fim:", estatisticas.get('end_time', '')],
        [""],
        ["1 - Sucesso:", sucessos],
        ["2 - Implantação ou PAB:", implantacao_ou_pab],
        ["3 - Falha na extração:", estatisticas.get('falha_extracao', 0)],
        ["4 - Falha no hiscre:", estatisticas.get('falha_hiscre', 0)],
        ["5 - Não Implantado:", estatisticas.get('nao_implantado', 0)],
        ["6 - Exige 2FA:", estatisticas.get('exige_2fa', 0)],
        ["7 - Senha Não Confere:", estatisticas.get('senha_errada', 0)],
        ["8 - Usuário Bloqueado:", estatisticas.get('usuario_bloqueado', 0)],
        ["9 - Senha Não Fornecida:", estatisticas.get('sem_senha', 0)],
        ["10 - Pdfs na pasta:", qtd_pdfs]
    ]

    for linha in linhas_resumo:
        ws_resumo.append(linha)
    
    # Regra de Cores do Resumo (Detalhe 6):
    # Se "Pdfs na pasta" == "Sucesso", a fonte fica verde, senão vermelha.
    # No nosso array acima, Implantação ou PAB tá na linha 6 e Pdfs na linha 13 (1-based index)
    cor_resumo_pdfs = fonte_verde if qtd_pdfs == implantacao_ou_pab else fonte_vermelha
    
    ws_resumo.cell(row=6, column=1).font = cor_resumo_pdfs
    ws_resumo.cell(row=6, column=2).font = cor_resumo_pdfs
    ws_resumo.cell(row=14, column=1).font = cor_resumo_pdfs
    ws_resumo.cell(row=14, column=2).font = cor_resumo_pdfs

    # ======== ABA 2: Consulta ========
    ws_consulta = wb.create_sheet(title="Consulta")
    
    colunas_header = [
        "CONSULTA", "MOTIVO", "AÇÃO", "CLIENTE", "CPF", "DATA", "VALOR", "BANCO", "PROCESSO", 
        "TIPO DE PROCESSO", "GRUPO DE PROCESSO", "ESFERA", "PARCEIRO", 
        "LÍDER", "PETICIONANTE", "DATA DE INCLUSÃO", "IMPLANTAÇÃO", "PAB"
    ]
    ws_consulta.append(colunas_header)
    
    # Estilizando o header da consulta
    for col in range(1, len(colunas_header) + 1):
        ws_consulta.cell(row=1, column=col).font = Font(bold=True)
    
    # Preenchendo dados
    data_hoje = date.today()
    
    for row_idx, dado in enumerate(dados_processados, start=2):
        
        # Trata o CPF para ser apenas números em formato de texto (evita formatação de milhar do excel)
        cpf_val = str(dado.get("CPF", "")).replace(".0", "")  # Se o pandas tiver lido como float
        for char in ".,- ":
            cpf_val = cpf_val.replace(char, "")
        cpf_val = cpf_val.zfill(11)  # Garante que tenha 11 dígitos, preenchendo com zeros à esquerda se necessário
            
        # Trata o PROCESSO para texto (remove ".0" caso tenha sido lido como número)
        processo_val = str(dado.get("PROCESSO", "")).replace(".0", "").strip()
        if processo_val == "nan" or processo_val == "None":
            processo_val = ""

        row_data = [
            dado.get("CONSULTA", ""),
            dado.get("MOTIVO", ""),
            dado.get("AÇÃO", ""),
            dado.get("CLIENTE", ""),
            cpf_val,
            dado.get("DATA", ""),  # Referente à "Previsão de pagamento" que vem do parser
            f'R$ {dado.get("VALOR", 0.0):.2f}'.replace('.', ','),
            dado.get("BANCO", ""),
            processo_val,
            dado.get("TIPO DE PROCESSO", ""),
            dado.get("GRUPO DE PROCESSO", ""),
            dado.get("ESFERA", ""),
            dado.get("PARCEIRO", ""),
            dado.get("LÍDER", ""),
            dado.get("PETICIONANTE", ""),
            dado.get("DATA DE INCLUSÃO", ""),
            dado.get("IMPLANTAÇÃO", ""),
            dado.get("PAB", "")
        ]
        ws_consulta.append(row_data)
        
        # --- APLICAR REGRAS DE CORES NAS LINHAS (Detalhes 3 e 4) ---
        cor_linha = fonte_padrao
        
        # Checa regra "Alerta de Implantação" (verde) - TEM PRIORIDADE ou podem coexistir? 
        # A regra 4 diz: Quando for alerta, linha toda verde.
        if dado.get("IMPLANTAÇÃO") == "ALERTA DE IMPLANTAÇÃO":
            cor_linha = fonte_verde
        else:
            # Checa regra "Data de Inclusão > 60 dias" (vermelha) (Detalhe 3)
            # Precisamos converter 'DATA DE INCLUSÃO' se for string. Assumindo formato DD/MM/YYYY.
            data_inc_str = dado.get("DATA DE INCLUSÃO", "")
            if isinstance(data_inc_str, datetime):
                data_inclusao_obj = data_inc_str.date()
            else:
                try:
                    # Tenta converter se for uma string real
                    if " " in data_inc_str: 
                        data_inc_str = data_inc_str.split(" ")[0] # limpar horas se houver
                    data_inclusao_obj = datetime.strptime(data_inc_str, "%d/%m/%Y").date()
                except:
                    data_inclusao_obj = None
            
            if data_inclusao_obj:
                diff_dias = (data_hoje - data_inclusao_obj).days
                if diff_dias > 60:
                    cor_linha = fonte_vermelha
                    
        # Aplica a cor em toda a linha
        for col_idx in range(1, len(colunas_header) + 1):
            ws_consulta.cell(row=row_idx, column=col_idx).font = cor_linha

    # Ajuste automático do tamanho das colunas (limite máximo de 50)
    for ws in [ws_resumo, ws_consulta]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        # Considera quebras de linha caso existam
                        linhas_celula = str(cell.value).split('\n')
                        for linha in linhas_celula:
                            max_length = max(max_length, len(linha))
                except:
                    pass
            
            # Adiciona um pequeno espaçamento (+2) e limita o tamanho a 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(caminho_excel)
    print(f"Relatório excel gerado com sucesso: {caminho_excel}")

    # Exportar também para ZIP incluindo o Excel + Pasta de PDFs
    nome_zip = f"{nome_excel.replace('.xlsx', '')}.zip"
    caminho_zip = os.path.join(pasta_destino_output, nome_zip)
    
    with zipfile.ZipFile(caminho_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Adiciona o excel
        zipf.write(caminho_excel, arcname=nome_excel)
        
        # Adiciona os PDFs
        for pdf_file in os.listdir(pasta_destino_pdfs):
            if pdf_file.endswith('.pdf'):
                caminho_pdf_interno = os.path.join(pasta_destino_pdfs, pdf_file)
                # Salva dentro de uma sub-pasta no zip chamada 'PDFs'
                zipf.write(caminho_pdf_interno, arcname=f"PDFs/{pdf_file}")
                
    print(f"Arquivo ZIP final gerado com sucesso: {caminho_zip}")
    return caminho_zip
